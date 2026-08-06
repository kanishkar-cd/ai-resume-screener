import csv
import json
from io import BytesIO, StringIO
from typing import Any

from app.core.exceptions import AppException


class ExportGenerationException(AppException):
    status_code = 500
    error_code = "EXPORT_GENERATION_FAILED"
    default_message = "Unable to generate the requested project export."


class ExportService:
    HEADERS = ("Rank", "Candidate Name", "Email", "Final Score", "Recommendation", "Confidence", "Skills Score", "Experience Score", "Matched Skills", "Missing Skills", "Status")

    @classmethod
    def generate_csv(cls, rows: list[dict[str, Any]]) -> bytes:
        try:
            output = StringIO(newline="")
            writer = csv.writer(output)
            writer.writerow(cls.HEADERS)
            for row in rows: writer.writerow(cls._values(row))
            return output.getvalue().encode("utf-8-sig")
        except Exception as exc: raise ExportGenerationException() from exc

    @classmethod
    def generate_excel(cls, rows: list[dict[str, Any]], analytics: dict[str, Any]) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.formatting.rule import ColorScaleRule
            from openpyxl.styles import Font, PatternFill
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Candidate Rankings"
            sheet.append(cls.HEADERS)
            for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78")
            for row in rows: sheet.append(cls._values(row))
            sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
            if sheet.max_row > 1: sheet.conditional_formatting.add(f"D2:D{sheet.max_row}", ColorScaleRule(start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B"))
            skills = workbook.create_sheet("Skills Analysis")
            skills.append(("Type", "Skill", "Frequency", "Percentage"))
            for kind, key in (("Matched", "top_matched_skills"), ("Missing", "top_missing_skills")):
                for item in analytics.get(key, []): skills.append((kind, item["skill_name"], item["frequency_count"], item["percentage"]))
            summary = workbook.create_sheet("Campaign Summary")
            for key in ("total_candidates", "average_score", "highest_score", "lowest_score", "average_confidence", "knocked_out_count"):
                summary.append((key.replace("_", " ").title(), analytics.get(key, 0)))
            output = BytesIO(); workbook.save(output); return output.getvalue()
        except Exception as exc: raise ExportGenerationException() from exc

    @classmethod
    def generate_pdf(cls, project: Any, rows: list[dict[str, Any]], analytics: dict[str, Any]) -> bytes:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
            output = BytesIO(); document = SimpleDocTemplate(output, pagesize=A4)
            styles = getSampleStyleSheet()
            story = [Paragraph(f"{project.title} - Candidate Report", styles["Title"]), Paragraph(f"Target role: {project.target_role}", styles["Normal"]), Spacer(1, 12)]
            summary = [["Candidates", "Average", "Highest", "Lowest", "Knocked Out"], [analytics.get("total_candidates", 0), analytics.get("average_score", 0), analytics.get("highest_score", 0), analytics.get("lowest_score", 0), analytics.get("knocked_out_count", 0)]]
            table = Table(summary); table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey)])); story.extend([table, Spacer(1, 16)])
            candidate_data = [["Rank", "Candidate", "Score", "Recommendation"]] + [[row["rank"], row["candidate_name"], row["final_score"], row["recommendation"]] for row in rows[:10]]
            candidates = Table(candidate_data, repeatRows=1); candidates.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), .5, colors.grey), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7"))])); story.append(candidates)
            document.build(story); return output.getvalue()
        except Exception as exc: raise ExportGenerationException() from exc

    @staticmethod
    def generate_json(project: Any, rows: list[dict[str, Any]], analytics: dict[str, Any]) -> bytes:
        try:
            payload = {"project": {"id": str(project.id), "title": project.title, "target_role": project.target_role}, "analytics": analytics, "candidates": rows}
            return json.dumps(payload, default=str, indent=2).encode("utf-8")
        except Exception as exc: raise ExportGenerationException() from exc

    @staticmethod
    def _values(row: dict[str, Any]) -> tuple[Any, ...]:
        return (row["rank"], row["candidate_name"], row.get("email") or "", row["final_score"], row["recommendation"], row["confidence"], row["skills_score"], row["experience_score"], ", ".join(row["matched_skills"]), ", ".join(row["missing_skills"]), "KNOCKED_OUT" if row["is_knocked_out"] else "ELIGIBLE")
