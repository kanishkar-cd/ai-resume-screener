import json
from types import SimpleNamespace
from uuid import uuid4

from app.services.export_service import ExportService


ROWS = [{"rank": 1, "candidate_name": "Jane Doe", "email": "jane@example.com", "final_score": 90, "recommendation": "STRONG_MATCH", "confidence": 95, "skills_score": 100, "experience_score": 80, "matched_skills": ["Python"], "missing_skills": ["Docker"], "is_knocked_out": False}]
ANALYTICS = {"total_candidates": 1, "average_score": 90, "highest_score": 90, "lowest_score": 90, "average_confidence": 95, "knocked_out_count": 0, "top_matched_skills": [{"skill_name": "Python", "frequency_count": 1, "percentage": 100}], "top_missing_skills": []}
PROJECT = SimpleNamespace(id=uuid4(), title="Backend Hiring", target_role="Engineer")


def test_csv_and_json_exports() -> None:
    csv_data = ExportService.generate_csv(ROWS)
    assert b"Jane Doe" in csv_data and csv_data.startswith(b"\xef\xbb\xbf")
    payload = json.loads(ExportService.generate_json(PROJECT, ROWS, ANALYTICS))
    assert payload["project"]["title"] == "Backend Hiring" and payload["candidates"][0]["rank"] == 1


def test_excel_export_is_valid_workbook() -> None:
    from openpyxl import load_workbook
    from io import BytesIO
    data = ExportService.generate_excel(ROWS, ANALYTICS)
    workbook = load_workbook(BytesIO(data))
    assert workbook.sheetnames == ["Candidate Rankings", "Skills Analysis", "Campaign Summary"]


def test_pdf_export_has_pdf_signature() -> None:
    data = ExportService.generate_pdf(PROJECT, ROWS, ANALYTICS)
    assert data.startswith(b"%PDF") and len(data) > 500
